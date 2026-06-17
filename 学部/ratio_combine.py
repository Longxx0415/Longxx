import pandas as pd
import numpy as np
from scipy.optimize import minimize
from sklearn.linear_model import Ridge
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import LeaveOneOut, cross_val_score
import matplotlib.pyplot as plt
import warnings
warnings.filterwarnings('ignore')
import matplotlib
matplotlib.use('TkAgg')  # 放在最前面，在 import matplotlib.pyplot as plt 之前

# ========== 1. 数据读取与对齐 ==========
df22 = pd.read_csv('data2022.csv', encoding='utf-8-sig')
df23 = pd.read_csv('data2023.csv', encoding='utf-8-sig')
df24 = pd.read_csv('data2024.csv', encoding='utf-8-sig')
df25 = pd.read_csv('data2025.csv', encoding='utf-8-sig')
df26 = pd.read_csv('data2026.csv', encoding='utf-8-sig')
df30 = pd.read_csv('forecastdata.csv', encoding='utf-8-sig')

college_names = df26['部门名称'].tolist()

def align_df(df, year):
    result = []
    for name in college_names:
        row = df[df['部门名称'] == name]
        if len(row) == 0:
            result.append({'部门名称': name, '年份': year})
        else:
            d = row.iloc[0].to_dict()
            d['年份'] = year
            result.append(d)
    return pd.DataFrame(result)

df22_a = align_df(df22, 2022)
df23_a = align_df(df23, 2023)
df24_a = align_df(df24, 2024)
df25_a = align_df(df25, 2025)
df26_a = align_df(df26, 2026)
df30_a = align_df(df30, 2030)

years = [2022, 2023, 2024, 2025, 2026]
hist_data = {y: d.set_index('部门名称') for y, d in zip(years, [df22_a, df23_a, df24_a, df25_a, df26_a])}
df30_idx = df30_a.set_index('部门名称')

# 读取OLS拟合值
from openpyxl import load_workbook
wb = load_workbook('测算数据（两个参数）.xlsx')
ws = wb['不含组织员']
ols_data = []
for row in ws.iter_rows(min_row=2, max_row=30, values_only=True):
    if row[1] is not None and row[1] != '学院':
        ols_data.append(row)
ols_df = pd.DataFrame(ols_data, columns=['序号', '学院', '管理岗现状', '学生数', '教师数', '拟合数', '吻合度'])
ols_fitted_dict = dict(zip(ols_df['学院'], pd.to_numeric(ols_df['拟合数'], errors='coerce')))
ols_fitted = np.array([ols_fitted_dict.get(name, np.nan) for name in college_names])

# ========== 2. 计算历史比例与稳定性判断 ==========
prop_data = {}
for year in years:
    total = hist_data[year]['编制人数'].sum()
    props = []
    for name in college_names:
        val = hist_data[year].loc[name, '编制人数']
        props.append(val / total if pd.notna(val) else np.nan)
    prop_data[year] = np.array(props)

prop_df = pd.DataFrame(prop_data, index=college_names)
prop_df['均值'] = prop_df.mean(axis=1)
prop_df['标准差'] = prop_df.std(axis=1)
prop_df['变异系数'] = prop_df['标准差'] / prop_df['均值']
prop_df['最大变化'] = prop_df[[2022, 2023, 2024, 2025, 2026]].max(axis=1) - prop_df[[2022, 2023, 2024, 2025, 2026]].min(axis=1)

stable_threshold_cv = 0.15
stable_threshold_maxchange = 0.015
stable_mask = (prop_df['变异系数'] < stable_threshold_cv) & (prop_df['最大变化'] < stable_threshold_maxchange)
stable_indices = [i for i in range(len(college_names)) if stable_mask.iloc[i]]
unstable_indices = [i for i in range(len(college_names)) if not stable_mask.iloc[i]]

print(f"稳定学院（{len(stable_indices)}个）: {[college_names[i] for i in stable_indices]}")
print(f"不稳定学院（{len(unstable_indices)}个）: {[college_names[i] for i in unstable_indices]}")

# ========== 3. 特征工程 ==========
feature_cols = ['本科生', '研究生', '在编-教师']

def build_features(df, name):
    row = df.loc[name]
    features = []
    for col in feature_cols:
        v = row[col] if col in df.columns and pd.notna(row[col]) else 0
        features.append(v)
    v = row['留学生人数'] if '留学生人数' in df.columns and pd.notna(row['留学生人数']) else 0
    features.append(v)
    v = row['公共课学分'] if '公共课学分' in df.columns and pd.notna(row['公共课学分']) else 0
    features.append(v)
    return features

# ========== 4. 方案一：稳定保持 + 不稳定Ridge预测 ==========
train_X, train_y = [], []
for year in years:
    for name in college_names:
        row = hist_data[year].loc[name]
        if pd.isna(row['编制人数']):
            continue
        train_X.append(build_features(hist_data[year], name))
        total = hist_data[year]['编制人数'].sum()
        train_y.append(row['编制人数'] / total)

scaler = StandardScaler()
train_X_scaled = scaler.fit_transform(np.array(train_X))
train_y = np.array(train_y)

best_alpha, best_score = None, -np.inf
for alpha in [0.001, 0.01, 0.1, 1, 10, 100, 1000]:
    model = Ridge(alpha=alpha)
    scores = cross_val_score(model, train_X_scaled, train_y, cv=LeaveOneOut(), scoring='neg_mean_squared_error')
    if scores.mean() > best_score:
        best_score, best_alpha = scores.mean(), alpha

ridge1 = Ridge(alpha=best_alpha)
ridge1.fit(train_X_scaled, train_y)

X30 = np.array([build_features(df30_idx, name) for name in college_names])
X30_scaled = scaler.transform(X30)
pred_props_ridge = np.maximum(ridge1.predict(X30_scaled), 0.001)

pred_props_1 = np.zeros(len(college_names))
for i in range(len(college_names)):
    pred_props_1[i] = prop_data[2026][i] if i in stable_indices else pred_props_ridge[i]
pred_props_1 = pred_props_1 / pred_props_1.sum()

# ========== 5. 约束优化 ==========
budget = np.array([df30_idx.loc[name, '预算约束人数'] if pd.notna(df30_idx.loc[name, '预算约束人数']) else 999 for name in college_names])
total_staff = 265

def optimize(target, budget, total):
    def objective(x): return np.sum((x - target)**2)
    def constraint(x): return np.sum(x) - total
    result = minimize(objective, target, method='SLSQP', bounds=[(0, b) for b in budget],
                      constraints={'type': 'eq', 'fun': constraint}, options={'maxiter': 1000})
    return result.x

pred_1 = optimize(pred_props_1 * total_staff, budget, total_staff)

# ========== 6. 与OLS比较 ==========
rmse1 = np.sqrt(np.mean((pred_1 - ols_fitted)**2))
mae1 = np.mean(np.abs(pred_1 - ols_fitted))

comparison = pd.DataFrame({
    '学院': college_names,
    '稳定': ['是' if i in stable_indices else '否' for i in range(len(college_names))],
    '方案一预测': np.round(pred_1, 4),
    'OLS拟合': np.round(ols_fitted, 4),
    '差异': np.round(pred_1 - ols_fitted, 4),
    '吻合度': np.round(np.minimum(pred_1, ols_fitted) / np.maximum(pred_1, ols_fitted), 4)
})
print("\n=== 方案一 vs OLS 数值对比 ===")
print(comparison.to_string(index=False))
print(f"\nRMSE={rmse1:.4f}, MAE={mae1:.4f}")

# ========== 7. 生成对比图（修复中文字体） ==========
plt.rcParams['font.sans-serif'] = ['SimHei', 'DejaVu Sans', 'Arial Unicode MS', 'Noto Sans CJK SC']
plt.rcParams['axes.unicode_minus'] = False

college_short = [
    '物理', '数学', '化学', '海洋', '生命',
    '机械', '测绘', '计算机', '汽车', '材料',
    '电子', '环境', '土木', '交通', '建筑',
    '航空', '中德', '医学', '设计', '人文',
    '外语', '经管', '政国', '艺传', '知识产权',
    '马院', '国交', '法学', '体育'
]

fig, ax = plt.subplots(figsize=(22, 12))

x = np.arange(len(college_names))
width = 0.35

bars1 = ax.bar(x - width/2, pred_1, width, label='方案一预测', color='#4472C4', alpha=0.85, edgecolor='navy', linewidth=0.5)
bars2 = ax.bar(x + width/2, ols_fitted, width, label='OLS拟合', color='#ED7D31', alpha=0.85, edgecolor='darkred', linewidth=0.5)

ax.set_xlabel('学院', fontsize=16, fontweight='bold')
ax.set_ylabel('编制人数', fontsize=16, fontweight='bold')
ax.set_title('方案一预测值 vs OLS拟合值对比（全部29个学院）', fontsize=18, fontweight='bold')
ax.set_xticks(x)
ax.set_xticklabels(college_short, rotation=45, ha='right', fontsize=12)
ax.legend(fontsize=14, loc='upper right')
ax.grid(True, alpha=0.3, axis='y', linestyle='--')

# 差异标注（差异>1.5的显示）
diff = pred_1 - ols_fitted
for i in range(len(college_names)):
    if abs(diff[i]) > 1.5:
        ax.text(i, max(pred_1[i], ols_fitted[i]) + 0.6, f'{diff[i]:+.1f}',
                ha='center', va='bottom', fontsize=10, color='red', fontweight='bold')

ax.set_ylim(0, max(max(pred_1), max(ols_fitted)) * 1.18)

# RMSE/MAE标注
ax.text(0.02, 0.98, f'RMSE = {rmse1:.4f}\nMAE = {mae1:.4f}', transform=ax.transAxes,
        fontsize=12, verticalalignment='top', bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.9))

plt.tight_layout()
plt.savefig('方案一_vs_OLS_对比图.png', dpi=300, bbox_inches='tight')
plt.show()
print("\n图片已保存: 方案一_vs_OLS_对比图.png")